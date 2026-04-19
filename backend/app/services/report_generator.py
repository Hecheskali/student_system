import io
import csv
from typing import Any, cast

from reportlab.lib import colors, pagesizes
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Flowable,
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    Image,
)
from reportlab.lib.enums import TA_LEFT
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.reports import (
    ReportExportDataSchema,
    ReportFormatEnum,
)


class ReportColors:
    """Professional color palette for reports"""

    PRIMARY = colors.HexColor("#1450B3")  # Professional blue
    PRIMARY_DARK = colors.HexColor("#051432")  # Dark blue
    ACCENT = colors.HexColor("#E65A26")  # Professional orange
    LIGHT_BG = colors.HexColor("#F5F7FC")  # Light blue-gray
    TEXT_DARK = colors.HexColor("#262D33")  # Dark text
    TEXT_LIGHT = colors.HexColor("#737C88")  # Light text
    BORDER = colors.HexColor("#CDD5E0")  # Border color


class ReportGenerator:
    """Generate professional formatted reports in various formats"""

    @staticmethod
    def generate(
        report_data: ReportExportDataSchema,
        file_format: ReportFormatEnum,
    ) -> tuple[bytes, str]:
        """
        Generate a report in the specified format.
        Returns tuple of (file_bytes, mime_type)
        """
        if file_format == ReportFormatEnum.PDF:
            return ReportGenerator._generate_pdf(report_data), "application/pdf"
        elif file_format == ReportFormatEnum.EXCEL:
            return ReportGenerator._generate_excel(report_data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif file_format == ReportFormatEnum.CSV:
            return ReportGenerator._generate_csv(report_data), "text/csv"
        else:
            raise ValueError(f"Unsupported report format: {file_format}")

    @staticmethod
    def _generate_pdf(report_data: ReportExportDataSchema) -> bytes:
        """Generate a professional PDF report"""
        buffer = io.BytesIO()

        # Determine page orientation
        page_size = pagesizes.landscape(
            pagesizes.A4
        ) if report_data.pdf_landscape else pagesizes.A4

        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            rightMargin=32,
            leftMargin=32,
            topMargin=24,
            bottomMargin=24,
        )

        # Build story
        story: list[Flowable] = []

        # Add header
        story.append(
            ReportGenerator._build_pdf_header(report_data)
        )
        story.append(Spacer(1, 12))

        # Add metadata
        if report_data.report_type or report_data.exam_window_label:
            story.append(
                ReportGenerator._build_pdf_metadata(report_data)
            )
            story.append(Spacer(1, 12))

        # Add summary boxes
        if report_data.summary:
            story.append(
                ReportGenerator._build_pdf_summary_boxes(report_data.summary)
            )
            story.append(Spacer(1, 12))

        # Add sections
        for i, section in enumerate(report_data.sections):
            story.append(
                ReportGenerator._build_pdf_section(section, report_data.pdf_landscape)
            )
            if i < len(report_data.sections) - 1:
                story.append(Spacer(1, 12))

        # Add footer
        if report_data.footnote:
            story.append(Spacer(1, 12))
            story.append(
                ReportGenerator._build_pdf_footer(report_data.footnote)
            )

        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _build_pdf_header(report_data: ReportExportDataSchema) -> Flowable:
        """Build professional PDF header"""
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=ReportColors.PRIMARY_DARK,
            spaceAfter=6,
            fontName="Helvetica-Bold",
            letterSpacing=0.5,
        )

        subtitle_style = ParagraphStyle(
            "CustomSubtitle",
            parent=styles["Normal"],
            fontSize=10,
            textColor=ReportColors.TEXT_LIGHT,
            spaceAfter=0,
            fontName="Helvetica",
        )

        content = f"<b>{report_data.title.upper()}</b>"
        header_para = Paragraph(content, title_style)

        if report_data.subtitle:
            subtitle_content = report_data.subtitle
            subtitle_para = Paragraph(subtitle_content, subtitle_style)
            from reportlab.platypus import Table as RLTable

            return RLTable(
                [[header_para], [subtitle_para]],
                colWidths=[7 * inch],
                style=TableStyle([("BOTTOMPADDING", (0, 0), (-1, 0), 6)]),
            )

        return header_para

    @staticmethod
    def _build_pdf_metadata(report_data: ReportExportDataSchema) -> Flowable:
        """Build PDF metadata section"""
        metadata_items = []

        if report_data.report_type:
            metadata_items.append(
                [
                    Paragraph(
                        "<b>Report Type</b>",
                        ParagraphStyle(
                            "MetaLabel",
                            fontSize=8,
                            textColor=ReportColors.TEXT_LIGHT,
                            fontName="Helvetica",
                        ),
                    ),
                    Paragraph(
                        report_data.report_type,
                        ParagraphStyle(
                            "MetaValue",
                            fontSize=10,
                            textColor=ReportColors.TEXT_DARK,
                            fontName="Helvetica-Bold",
                        ),
                    ),
                ]
            )

        if report_data.exam_window_label:
            metadata_items.append(
                [
                    Paragraph(
                        "<b>Exam Period</b>",
                        ParagraphStyle(
                            "MetaLabel",
                            fontSize=8,
                            textColor=ReportColors.TEXT_LIGHT,
                            fontName="Helvetica",
                        ),
                    ),
                    Paragraph(
                        report_data.exam_window_label,
                        ParagraphStyle(
                            "MetaValue",
                            fontSize=10,
                            textColor=ReportColors.TEXT_DARK,
                            fontName="Helvetica-Bold",
                        ),
                    ),
                ]
            )

        if not metadata_items:
            return Spacer(1, 0)

        return Table(
            metadata_items,
            colWidths=[1.5 * inch, 4 * inch],
            style=TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]),
        )

    @staticmethod
    def _build_pdf_summary_boxes(summary_items: list[Any]) -> Table:
        """Build summary information boxes"""
        rows = []
        cols = 3
        for i in range(0, len(summary_items), cols):
            row = []
            for item in summary_items[i : i + cols]:
                label_para = Paragraph(
                    item.label,
                    ParagraphStyle(
                        "SummaryLabel",
                        fontSize=8,
                        textColor=ReportColors.TEXT_LIGHT,
                        fontName="Helvetica-Bold",
                    ),
                )
                value_para = Paragraph(
                    item.value,
                    ParagraphStyle(
                        "SummaryValue",
                        fontSize=11,
                        textColor=ReportColors.PRIMARY,
                        fontName="Helvetica-Bold",
                    ),
                )
                box = Table(
                    [[label_para], [value_para]],
                    colWidths=[1.3 * inch],
                    style=TableStyle([
                        ("BACKGROUND", (0, 0), (-1, -1), ReportColors.LIGHT_BG),
                        ("LEFTPADDING", (0, 0), (-1, -1), 10),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                        ("BORDER", (0, 0), (-1, -1), 1, ReportColors.BORDER),
                    ]),
                )
                row.append(box)

            rows.append(row)

        return Table(
            rows,
            colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch],
            style=TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]),
        )

    @staticmethod
    def _build_pdf_section(section_data: Any, pdf_landscape: bool) -> Flowable:
        """Build a data table section"""
        del pdf_landscape
        title = Paragraph(
            section_data.title,
            ParagraphStyle(
                "SectionTitle",
                fontSize=12,
                textColor=ReportColors.PRIMARY_DARK,
                fontName="Helvetica-Bold",
                spaceAfter=6,
                letterSpacing=0.3,
            ),
        )

        # Create data rows
        table_data = [section_data.headers]
        for row in section_data.rows:
            table_data.append([str(cell) if cell is not None else "" for cell in row])

        # Create table with professional styling
        col_count = len(section_data.headers)
        col_width = (6.5 * inch) / col_count if col_count > 0 else 1 * inch

        if section_data.pdf_column_flexes:
            total_flex = sum(section_data.pdf_column_flexes)
            col_widths = [
                (6.5 * inch) * (flex / total_flex)
                for flex in section_data.pdf_column_flexes
            ]
        else:
            col_widths = [col_width] * col_count

        table = Table(
            table_data,
            colWidths=col_widths,
            repeatRows=1,
        )

        # Style table
        styles = [
            ("BACKGROUND", (0, 0), (-1, 0), ReportColors.PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("TEXTCOLOR", (0, 1), (-1, -1), ReportColors.TEXT_DARK),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ReportColors.LIGHT_BG]),
            ("GRID", (0, 0), (-1, -1), 0.5, ReportColors.BORDER),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("TOPPADDING", (0, 1), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]

        table.setStyle(TableStyle(styles))

        return Table(
            [[title], [table]],
            colWidths=[7 * inch],
            style=TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0)]),
        )

    @staticmethod
    def _build_pdf_footer(footnote: str) -> Paragraph:
        """Build PDF footer"""
        return Paragraph(
            footnote,
            ParagraphStyle(
                "Footer",
                fontSize=8,
                textColor=ReportColors.TEXT_LIGHT,
                fontName="Helvetica",
                alignment=TA_LEFT,
            ),
        )

    @staticmethod
    def _generate_excel(report_data: ReportExportDataSchema) -> bytes:
        """Generate an Excel report"""
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(title="Report")
        ws = _coerce_worksheet(ws)
        ws.title = "Report"

        current_row = 1

        # Add title
        title_cell = _worksheet_cell(ws, current_row, 1)
        title_cell.value = report_data.title
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="1450B3")
        ws.merge_cells(f"A{current_row}:D{current_row}")
        current_row += 2

        # Add metadata
        if report_data.school_name:
            ws[f"A{current_row}"] = "School"
            ws[f"B{current_row}"] = report_data.school_name
            current_row += 1

        if report_data.report_type:
            ws[f"A{current_row}"] = "Report Type"
            ws[f"B{current_row}"] = report_data.report_type
            current_row += 1

        if report_data.exam_window_label:
            ws[f"A{current_row}"] = "Exam Period"
            ws[f"B{current_row}"] = report_data.exam_window_label
            current_row += 1

        if report_data.generated_at:
            ws[f"A{current_row}"] = "Generated"
            ws[f"B{current_row}"] = report_data.generated_at.isoformat()
            current_row += 1

        current_row += 1

        # Add sections
        for section in report_data.sections:
            # Section title
            title_cell = _worksheet_cell(ws, current_row, 1)
            title_cell.value = section.title
            title_cell.font = Font(name="Calibri", size=12, bold=True, color="051432")
            ws.merge_cells(f"A{current_row}:Z{current_row}")
            current_row += 1

            # Headers
            header_fill = PatternFill(
                start_color="1450B3", end_color="1450B3", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, header in enumerate(section.headers, start=1):
                cell = _worksheet_cell(ws, current_row, col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            current_row += 1

            # Data rows
            border = Border(
                left=Side(style="thin", color="CDD5E0"),
                right=Side(style="thin", color="CDD5E0"),
                top=Side(style="thin", color="CDD5E0"),
                bottom=Side(style="thin", color="CDD5E0"),
            )

            row_fill_colors = [None, PatternFill(start_color="F5F7FC", end_color="F5F7FC", fill_type="solid")]

            for row_idx, row_data in enumerate(section.rows):
                fill = row_fill_colors[row_idx % 2]
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = _worksheet_cell(ws, current_row, col_idx)
                    cell.value = cell_value
                    cell.border = border
                    if fill:
                        cell.fill = fill
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                current_row += 1

            current_row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _generate_csv(report_data: ReportExportDataSchema) -> bytes:
        """Generate a CSV report"""
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # Write title
        writer.writerow([report_data.title])
        writer.writerow([])

        # Write metadata
        if report_data.school_name:
            writer.writerow(["School", report_data.school_name])

        if report_data.report_type:
            writer.writerow(["Report Type", report_data.report_type])

        if report_data.exam_window_label:
            writer.writerow(["Exam Period", report_data.exam_window_label])

        if report_data.generated_at:
            writer.writerow(["Generated", report_data.generated_at.isoformat()])

        writer.writerow([])

        # Write sections
        for section in report_data.sections:
            writer.writerow([section.title])
            if section.note:
                writer.writerow([section.note])

            writer.writerow(section.headers)
            for row in section.rows:
                writer.writerow([cell if cell is not None else "" for cell in row])

            writer.writerow([])

        if report_data.footnote:
            writer.writerow([])
            writer.writerow([report_data.footnote])

        return buffer.getvalue().encode("utf-8")


def _coerce_worksheet(worksheet: Any) -> Worksheet:
    if not isinstance(worksheet, Worksheet):
        raise TypeError("Expected an openpyxl worksheet")
    return worksheet


def _worksheet_cell(worksheet: Worksheet, row: int, column: int) -> Cell:
    return cast(Cell, worksheet.cell(row=row, column=column))
